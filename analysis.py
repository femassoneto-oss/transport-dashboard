import os
from collections import defaultdict, Counter
from datetime import datetime
from openpyxl import load_workbook
import itertools

# ---------------------------------------------------------------------------
# Helper to read the Excel file into a list of dictionaries
# ---------------------------------------------------------------------------
EXCEL_PATH = r"Z:\1. Biomassa & Nutrição Animal\BRUNA - Copia\ANTIGRAVITY\data(51).xlsx"
SHEET_NAME = None  # read the first sheet if None


def _load_data():
    """Return a list of rows where each row is a dict with canonical keys.
    Expected columns (after mapping):
        - plate (str)
        - operation (str)
        - client (str)
        - date (datetime)
        - revenue (float)
        - productivity (float)
    """
    wb = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    # Assume first row contains headers
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)]
    # Normalise header names to the canonical ones (case‑insensitive match)
    mapping = {
        "Placa": "plate",
        "Operação": "operation",
        "Cliente": "client",
        "Data": "date",
        "Faturamento": "revenue",
        "Produtividade": "productivity",
    }
    # Build index map from column index to canonical key
    col_map = {}
    for idx, hdr in enumerate(headers):
        if hdr in mapping:
            col_map[idx] = mapping[hdr]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = {}
        for idx, key in col_map.items():
            val = row[idx]
            if key == "date" and isinstance(val, (datetime, )):
                entry[key] = val
            elif key in ("revenue", "productivity"):
                # Convert possible None or string to float
                try:
                    entry[key] = float(val) if val is not None else 0.0
                except Exception:
                    entry[key] = 0.0
            else:
                entry[key] = str(val).strip() if val is not None else ""
        # Skip rows missing essential fields
        if entry.get("plate") and entry.get("operation") and entry.get("date"):
            rows.append(entry)
    return rows

# ---------------------------------------------------------------------------
# Metric calculations (all work on the list of dicts returned by _load_data)
# ---------------------------------------------------------------------------

def total_plates(data):
    return len({row["plate"] for row in data})

def plates_per_operation(data):
    counter = Counter(row["operation"] for row in data)
    # Count unique plates per operation
    op_to_plates = defaultdict(set)
    for row in data:
        op_to_plates[row["operation"]].add(row["plate"])
    return {op: len(plates) for op, plates in op_to_plates.items()}

def unique_plates_per_client(data):
    client_to_plates = defaultdict(set)
    for row in data:
        client_to_plates[row["client"]].add(row["plate"])
    return {client: len(plates) for client, plates in client_to_plates.items()}

def plates_march_not_april(data):
    march = {row["plate"] for row in data if row["date"].month == 3}
    april = {row["plate"] for row in data if row["date"].month == 4}
    return list(march - april)

def shared_plates_between_operations(data):
    plate_ops = defaultdict(set)
    for row in data:
        plate_ops[row["plate"]].add(row["operation"])
    # Keep plates that appear in more than one operation
    shared = {plate: ops for plate, ops in plate_ops.items() if len(ops) > 1}
    # Return a list of tuples (plate, operation) for easier display
    result = []
    for plate, ops in shared.items():
        for op in ops:
            result.append((plate, op))
    return result

def revenue_per_operation(data):
    op_rev = defaultdict(float)
    for row in data:
        op_rev[row["operation"]] += row.get("revenue", 0.0)
    return dict(op_rev)

def productivity_per_operation(data):
    op_prod = defaultdict(float)
    for row in data:
        op_prod[row["operation"]] += row.get("productivity", 0.0)
    return dict(op_prod)

def plate_ranking(data):
    plate_stats = defaultdict(lambda: {"revenue": 0.0, "productivity": 0.0})
    for row in data:
        p = row["plate"]
        plate_stats[p]["revenue"] += row.get("revenue", 0.0)
        plate_stats[p]["productivity"] += row.get("productivity", 0.0)
    # Build list sorted by revenue descending
    ranking = []
    for plate, stats in plate_stats.items():
        ranking.append({"plate": plate, "total_revenue": stats["revenue"], "total_productivity": stats["productivity"]})
    ranking.sort(key=lambda x: x["total_revenue"], reverse=True)
    for i, item in enumerate(ranking, start=1):
        item["rank"] = i
    return ranking

def growth_operations(data):
    # month‑over‑month revenue per operation
    op_month_rev = defaultdict(lambda: defaultdict(float))
    for row in data:
        month = row["date"].replace(day=1)  # first day of month as key
        op_month_rev[row["operation"]][month] += row.get("revenue", 0.0)
    growth_list = []
    for op, month_dict in op_month_rev.items():
        # sort months chronologically
        months = sorted(month_dict.keys())
        for prev, cur in zip(months, months[1:]):
            prev_rev = month_dict[prev]
            cur_rev = month_dict[cur]
            if prev_rev == 0:
                continue
            growth_pct = (cur_rev - prev_rev) / prev_rev * 100.0
            growth_list.append({
                "operation": op,
                "prev_month": prev.strftime("%Y-%m"),
                "cur_month": cur.strftime("%Y-%m"),
                "growth_pct": growth_pct,
            })
    # Return the best growth per operation (largest positive pct)
    best = {}
    for item in growth_list:
        op = item["operation"]
        if op not in best or item["growth_pct"] > best[op]["growth_pct"]:
            best[op] = item
    return list(best.values())

def top_clients(data):
    client_stats = defaultdict(lambda: {"revenue": 0.0, "plates": set()})
    for row in data:
        c = row["client"]
        client_stats[c]["revenue"] += row.get("revenue", 0.0)
        client_stats[c]["plates"].add(row["plate"])
    result = []
    for client, stats in client_stats.items():
        result.append({
            "client": client,
            "total_revenue": stats["revenue"],
            "unique_plates": len(stats["plates"])},
        )
    # Composite score (70% revenue, 30% plates) for ranking
    max_rev = max((r["total_revenue"] for r in result), default=1)
    max_plates = max((r["unique_plates"] for r in result), default=1)
    for r in result:
        r["score"] = 0.7 * (r["total_revenue"] / max_rev) + 0.3 * (r["unique_plates"] / max_plates)
    result.sort(key=lambda x: x["score"], reverse=True)
    return result

# ---------------------------------------------------------------------------
# Simple automated insights / anomaly detection
# ---------------------------------------------------------------------------

def generate_insights(data):
    insights = []
    # 1️⃣ Revenue drop >20% month‑over‑month
    month_rev = defaultdict(float)
    for row in data:
        month = row["date"].replace(day=1)
        month_rev[month] += row.get("revenue", 0.0)
    sorted_months = sorted(month_rev.keys())
    for prev, cur in zip(sorted_months, sorted_months[1:]):
        prev_rev = month_rev[prev]
        cur_rev = month_rev[cur]
        if prev_rev == 0:
            continue
        pct_change = (cur_rev - prev_rev) / prev_rev * 100.0
        if pct_change < -20:
            insights.append(f"⚠️ Receita caiu {abs(pct_change):.1f}% de {prev.strftime('%B %Y')} para {cur.strftime('%B %Y')}.")
    # 2️⃣ Productivity outliers (z‑score > 2.5)
    prod_vals = [row.get("productivity", 0.0) for row in data]
    if prod_vals:
        mean = sum(prod_vals) / len(prod_vals)
        var = sum((x - mean) ** 2 for x in prod_vals) / len(prod_vals)
        std = var ** 0.5
        outliers = [row for row in data if std and abs(row.get("productivity", 0.0) - mean) / std > 2.5]
        if outliers:
            insights.append(f"🔎 Detectados {len(outliers)} registros com produtividade fora do padrão.")
    # 3️⃣ Placas que operaram em março mas não em abril
    diff = plates_march_not_april(data)
    if diff:
        insights.append(f"ℹ️ {len(diff)} placas operaram em março e não em abril.")
    return insights

# ---------------------------------------------------------------------------
# Public API – a single function that returns everything ready for the UI
# ---------------------------------------------------------------------------

def get_dashboard_data():
    data = _load_data()
    return {
        "raw": data,
        "total_plates": total_plates(data),
        "plates_per_operation": plates_per_operation(data),
        "unique_plates_per_client": unique_plates_per_client(data),
        "march_not_april": plates_march_not_april(data),
        "shared_plates": shared_plates_between_operations(data),
        "revenue_per_operation": revenue_per_operation(data),
        "productivity_per_operation": productivity_per_operation(data),
        "plate_ranking": plate_ranking(data),
        "growth_operations": growth_operations(data),
        "top_clients": top_clients(data),
        "insights": generate_insights(data),
    }

if __name__ == "__main__":
    # Simple CLI test – prints a summary to console
    from pprint import pprint
    pprint(get_dashboard_data())
